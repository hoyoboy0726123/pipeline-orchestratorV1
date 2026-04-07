"""
Pipeline 狀態機主引擎。

流程：
  START → 逐步執行 → LLM 驗證 → 通過則下一步
                                 → 失敗且有重試次數 → 自動重試
                                 → 失敗且重試耗盡  → 暫停 + Telegram inline keyboard
  用戶按 [重試 / 跳過 / 中止] → resume_pipeline() 繼續或結束

Telegram 通知時機：
  - 步驟失敗需人為決策 → 詢問訊息 + inline keyboard
  - Pipeline 全部完成 / 中止 → 結果摘要
"""
import asyncio
import logging
import uuid
from datetime import datetime
from typing import Optional

from telegram import Bot, InlineKeyboardButton, InlineKeyboardMarkup

from config import TELEGRAM_BOT_TOKEN
from .models import PipelineConfig
from .store import PipelineRun, StepResult, get_store
from .logger import create_run_logger
from .executor import execute_step, execute_step_with_skill
from .validator import validate_step, validate_step_with_skill, ValidationResult


# ── Telegram helpers ─────────────────────────────────────────────────────────

def _decision_keyboard(run_id: str) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([[
        InlineKeyboardButton("🔄 重試此步驟", callback_data=f"pipe_retry:{run_id}"),
        InlineKeyboardButton("⏩ 跳過此步驟", callback_data=f"pipe_skip:{run_id}"),
        InlineKeyboardButton("🛑 中止",       callback_data=f"pipe_abort:{run_id}"),
    ]])


async def _tg_send(chat_id: int, text: str, reply_markup=None):
    """發送 Telegram 訊息（錯誤靜默記錄，不拋出）"""
    if not chat_id or not TELEGRAM_BOT_TOKEN:
        return
    try:
        bot = Bot(token=TELEGRAM_BOT_TOKEN)
        await bot.send_message(
            chat_id=chat_id,
            text=text,
            parse_mode="HTML",
            reply_markup=reply_markup,
        )
        await bot.close()
    except Exception as e:
        logging.getLogger("pipeline").error(f"Telegram 發送失敗：{e}")


async def _notify_failure(run: PipelineRun, val: ValidationResult, step_name: str):
    """詢問用戶如何處理失敗步驟"""
    step_num = run.current_step + 1
    total = len(PipelineConfig.from_dict(run.config_dict).steps)
    text = (
        f"⚠️ <b>Pipeline 需要決策</b>\n\n"
        f"📋 {run.pipeline_name}\n"
        f"📍 步驟 {step_num}/{total}：<b>{step_name}</b>\n\n"
        f"🔴 {val.reason}\n"
    )
    if val.suggestion:
        text += f"💡 建議：{val.suggestion}\n"
    text += "\n請選擇處理方式："
    await _tg_send(run.telegram_chat_id, text, _decision_keyboard(run.run_id))


async def _notify_final(run: PipelineRun, config: PipelineConfig):
    """發送 pipeline 最終結果摘要"""
    total = len(config.steps)
    ok_count = sum(1 for r in run.step_results if r.validation_status == "ok")

    status_map = {
        "completed": ("✅", "Pipeline 完成"),
        "aborted":   ("🛑", "Pipeline 已中止"),
    }
    emoji, title = status_map.get(run.status, ("❌", "Pipeline 失敗"))

    duration = ""
    if run.ended_at and run.started_at:
        try:
            secs = int((
                datetime.fromisoformat(run.ended_at) -
                datetime.fromisoformat(run.started_at)
            ).total_seconds())
            duration = f"⏱ 耗時：{secs // 60}m {secs % 60}s\n"
        except Exception:
            pass

    # Step 摘要
    step_lines = []
    for i, step in enumerate(config.steps):
        if i < len(run.step_results):
            r = run.step_results[i]
            icon = {"ok": "✅", "warning": "⚠️", "failed": "❌"}.get(r.validation_status, "❓")
            step_lines.append(f"  {icon} {step.name}")
        else:
            step_lines.append(f"  ⬜ {step.name}（未執行）")

    text = (
        f"{emoji} <b>{title}</b>\n\n"
        f"📋 {run.pipeline_name}\n"
        f"🔢 {ok_count}/{total} 步驟成功\n"
        f"{duration}"
        f"\n<b>步驟概覽：</b>\n" + "\n".join(step_lines) +
        f"\n\n📁 <code>{run.log_path}</code>"
    )
    await _tg_send(run.telegram_chat_id, text)


# ── Deterministic validation (fast recipe mode) ──────────────────────────────

def _deterministic_validate(step, exec_result, logger) -> ValidationResult:
    """Recipe 快速模式：不叫 LLM，只做確定性檢查。"""
    from pathlib import Path as _Path

    # 1. exit code
    if exec_result.exit_code != 0:
        return ValidationResult(
            status="failed",
            reason=f"Exit code {exec_result.exit_code}",
            suggestion="Recipe 執行失敗，建議改用完整模式重跑",
        )

    # 2. 輸出檔存在 + 大小
    if step.output and step.output.path:
        p = _Path(step.output.path)
        if not p.exists():
            return ValidationResult(
                status="failed",
                reason=f"輸出檔案 {step.output.path} 不存在",
                suggestion="Recipe 未產生預期檔案，建議改用完整模式",
            )
        size = p.stat().st_size
        if size == 0:
            return ValidationResult(
                status="failed",
                reason=f"輸出檔案 {step.output.path} 為空檔案（0 bytes）",
                suggestion="Recipe 產生了空檔案，建議改用完整模式",
            )
        # CSV: 檢查有 header
        if p.suffix.lower() == ".csv":
            try:
                with open(p, "r", encoding="utf-8") as f:
                    lines = sum(1 for _ in f)
                if lines < 2:
                    return ValidationResult(
                        status="failed",
                        reason=f"CSV 檔案只有 {lines} 行（預期至少有 header + 資料）",
                        suggestion="",
                    )
            except Exception:
                pass
        # Excel: 檢查有 sheet
        if p.suffix.lower() in (".xlsx", ".xls"):
            try:
                from openpyxl import load_workbook
                wb = load_workbook(p, read_only=True)
                sheet_count = len(wb.sheetnames)
                wb.close()
                if sheet_count == 0:
                    return ValidationResult(
                        status="failed",
                        reason="Excel 檔案沒有任何工作表",
                        suggestion="",
                    )
            except Exception:
                pass

    logger.info(f"[{step.name}] ⚡ Recipe 快速驗證通過（確定性檢查）")
    return ValidationResult(
        status="ok",
        reason="Recipe 快速模式：exit code=0、輸出檔案存在且非空",
        suggestion="",
    )


# ── Main pipeline engine ──────────────────────────────────────────────────────

async def run_pipeline(
    config_dict: dict,
    chat_id: int,
    run_id: Optional[str] = None,
    start_from_step: int = 0,
) -> str:
    """
    執行（或恢復）一個 pipeline。

    Args:
        config_dict:     PipelineConfig dict
        chat_id:         Telegram chat id（用於通知）
        run_id:          None = 新建；有值 = 恢復現有 run
        start_from_step: 從哪一步開始（resume 用）

    Returns:
        run_id
    """
    store = get_store()

    # 建立或恢復 run
    if run_id:
        run = store.load(run_id)
        if not run:
            raise ValueError(f"找不到 pipeline run: {run_id}")
        run.status = "running"
        run.current_step = start_from_step
        logger, _ = create_run_logger(run.run_id, run.pipeline_name)
        logger.info(f"恢復執行，從步驟 {start_from_step + 1} 繼續")
    else:
        config = PipelineConfig.from_dict(config_dict)
        run_id = str(uuid.uuid4())[:12]
        logger, log_path = create_run_logger(run_id, config.name)
        run = PipelineRun(
            run_id=run_id,
            pipeline_name=config.name,
            config_dict=config_dict,
            telegram_chat_id=chat_id,
            log_path=log_path,
        )
        logger.info(f"Pipeline 開始：{config.name}，共 {len(config.steps)} 步驟")

    config = PipelineConfig.from_dict(run.config_dict)
    use_recipe = run.config_dict.get("_use_recipe", False)
    workflow_id = run.config_dict.get("_workflow_id") or run.workflow_id
    store.save(run)

    # ── Step loop ────────────────────────────────────────────
    completed_outputs: list[dict] = []  # 收集前步驟的輸出資訊

    while run.current_step < len(config.steps):
        step = config.steps[run.current_step]
        step_num = run.current_step + 1
        total = len(config.steps)
        logger.info(f"══ 步驟 {step_num}/{total}：{step.name} ══")
        logger.debug(f"[{step.name}] batch 全文（{len(step.batch)} 字元）：{step.batch[:500]}")

        retries_used = 0

        # Retry loop for this step
        while True:
            if step.skill_mode:
                # working_dir: 優先用 step 指定，fallback 到 output_path 的目錄
                wd = step.working_dir
                if not wd and step.output and step.output.path:
                    from pathlib import Path as _Path
                    wd = str(_Path(step.output.path).parent)
                exec_result = await execute_step_with_skill(
                    task_description=step.batch,
                    timeout=step.timeout,
                    logger=logger,
                    step_name=step.name,
                    output_path=step.output.path if step.output else None,
                    working_dir=wd or None,
                    prev_outputs=completed_outputs if completed_outputs else None,
                    pipeline_id=workflow_id or config.name,
                )
            else:
                exec_result = await execute_step(
                    command=step.batch,
                    timeout=step.timeout,
                    logger=logger,
                    step_name=step.name,
                )

            # 快速模式：Recipe 命中 + 執行成功 → 確定性驗證（不叫 LLM）
            recipe_hit = (exec_result.stderr == "__RECIPE_HIT__")
            if recipe_hit:
                exec_result.stderr = ""  # 清掉標記

            if recipe_hit and use_recipe and exec_result.exit_code == 0:
                # 確定性檢查：exit code=0、輸出檔存在、檔案大小合理
                val = _deterministic_validate(step, exec_result, logger)
            elif config.validate:
                # 完整 LLM 驗證
                use_skill = step.output and step.output.skill_mode
                validate_fn = validate_step_with_skill if use_skill else validate_step
                val = await validate_fn(
                    step_name=step.name,
                    command=step.batch,
                    exit_code=exec_result.exit_code,
                    stdout=exec_result.stdout,
                    stderr=exec_result.stderr,
                    output_path=step.output.path if step.output else None,
                    output_expect=step.output.get_expect() if step.output else None,
                    logger=logger,
                )
            else:
                status = "ok" if exec_result.exit_code == 0 else "failed"
                val = ValidationResult(
                    status=status,
                    reason=f"Exit code {exec_result.exit_code}（LLM 驗證已停用）",
                    suggestion="" if status == "ok" else "請查看 log 取得詳細錯誤",
                )
                logger.info(f"[{step.name}] 驗證（僅 exit code）：{val.status}")

            step_result = StepResult(
                step_index=run.current_step,
                step_name=step.name,
                exit_code=exec_result.exit_code,
                stdout_tail=exec_result.stdout[-500:],
                stderr_tail=exec_result.stderr[-200:],
                validation_status=val.status,
                validation_reason=val.reason,
                validation_suggestion=val.suggestion,
                retries_used=retries_used,
            )

            # 更新或追加步驟結果
            if len(run.step_results) > run.current_step:
                run.step_results[run.current_step] = step_result
            else:
                run.step_results.append(step_result)
            store.save(run)

            if val.status == "ok":
                logger.info(f"步驟 {step_num} ✅ 通過")
                # 收集此步驟的輸出資訊供後續步驟參考
                if step.output and step.output.path:
                    out_info = {"path": step.output.path, "schema": ""}
                    try:
                        from pathlib import Path as _Path
                        p = _Path(step.output.path)
                        if p.suffix == ".csv" and p.exists():
                            with open(p, "r") as f:
                                header = f.readline().strip()
                            out_info["schema"] = header
                        elif p.suffix in (".xlsx", ".xls") and p.exists():
                            out_info["schema"] = "Excel 工作簿"
                        elif p.suffix in (".png", ".jpg", ".jpeg") and p.exists():
                            out_info["schema"] = "圖片檔案"
                    except Exception:
                        pass
                    completed_outputs.append(out_info)
                run.current_step += 1
                store.save(run)
                break  # 進入下一步

            elif retries_used < step.retry:
                retries_used += 1
                logger.warning(
                    f"步驟 {step_num} 驗證失敗，自動重試 {retries_used}/{step.retry}：{val.reason}"
                )
                continue  # 重試

            else:
                # 重試耗盡，暫停等待人為決策
                logger.warning(f"步驟 {step_num} 失敗且重試次數耗盡，等待人為決策")
                run.status = "awaiting_human"
                store.save(run)
                await _notify_failure(run, val, step.name)
                return run.run_id  # 暫停

    # ── 全部步驟完成 ─────────────────────────────────────────
    run.status = "completed"
    run.ended_at = datetime.now().isoformat()
    store.save(run)
    logger.info(f"Pipeline {config.name} 全部完成！")
    await _notify_final(run, config)
    return run.run_id


# ── Human-in-the-loop resume ─────────────────────────────────────────────────

async def resume_pipeline(run_id: str, decision: str) -> str:
    """
    用戶透過 Telegram inline keyboard 做出決策後，呼叫此函式繼續執行。

    Args:
        run_id:   pipeline run id
        decision: "retry" | "skip" | "abort"

    Returns:
        str 回應訊息（回覆給用戶）
    """
    store = get_store()
    run = store.load(run_id)

    if not run:
        return f"❌ 找不到 Pipeline run：{run_id}"
    if run.status != "awaiting_human":
        return f"⚠️ Pipeline {run_id} 目前狀態為 {run.status}，無需決策"

    config = PipelineConfig.from_dict(run.config_dict)
    step_num = run.current_step + 1
    total = len(config.steps)

    if decision == "abort":
        run.status = "aborted"
        run.ended_at = datetime.now().isoformat()
        store.save(run)
        logger = logging.getLogger(f"pipeline.{run_id}")
        logger.info("用戶選擇中止 Pipeline")
        await _notify_final(run, config)
        return f"🛑 Pipeline 已中止（步驟 {step_num}/{total}）"

    elif decision == "skip":
        logger = logging.getLogger(f"pipeline.{run_id}")
        logger.info(f"用戶選擇跳過步驟 {step_num}")
        next_step = run.current_step + 1

        if next_step >= total:
            run.status = "completed"
            run.ended_at = datetime.now().isoformat()
            store.save(run)
            await _notify_final(run, config)
            return f"⏩ 跳過最後一步，Pipeline 完成"

        asyncio.create_task(run_pipeline(
            config_dict=run.config_dict,
            chat_id=run.telegram_chat_id,
            run_id=run.run_id,
            start_from_step=next_step,
        ))
        return f"⏩ 跳過步驟 {step_num}，繼續執行步驟 {step_num + 1}/{total}"

    elif decision == "retry":
        logger = logging.getLogger(f"pipeline.{run_id}")
        logger.info(f"用戶選擇重試步驟 {step_num}")
        asyncio.create_task(run_pipeline(
            config_dict=run.config_dict,
            chat_id=run.telegram_chat_id,
            run_id=run.run_id,
            start_from_step=run.current_step,
        ))
        return f"🔄 重試步驟 {step_num}/{total}"

    return "❓ 未知決策"
